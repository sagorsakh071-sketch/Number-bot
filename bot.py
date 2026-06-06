import asyncio
import re
import os
import json
import logging
import urllib.request
from datetime import datetime, date

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, filters, ContextTypes
)

# ─── Configuration ───
BOT_TOKEN      = "8609593081:AAEZczxKQZ4hPZ3wBuxp7zr_UFstUztMAAw"
OWNER_ID       = 7095358778
SUPPORT_USER   = "@sadhin8miya"
BAILEYS_URL    = os.environ.get("BAILEYS_URL", "http://localhost:8080")
ADMIN_WA_UID   = "wa_checker_admin"
DATA_FILE      = "checker_data.json"
USERS_PER_PAGE = 10

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════
# ─── Data Management ───
# ═══════════════════════════════════════

def load_data() -> dict:
    try:
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    except:
        return {
            "users": {},
            "banned": [],
            "approved": [],
            "global_wa": {"enabled": True, "connected": False},
            "settings": {"daily_limit": 0, "open_mode": True},
        }

def save_data(d: dict):
    with open(DATA_FILE, "w") as f:
        json.dump(d, f, indent=2, ensure_ascii=False)

db = load_data()

def get_user(uid: str) -> dict:
    return db["users"].get(uid, {})

def register_user(update: Update):
    uid   = str(update.effective_user.id)
    name  = update.effective_user.full_name
    uname = update.effective_user.username or ""
    if uid not in db["users"]:
        db["users"][uid] = {
            "name": name, "username": uname,
            "joined": datetime.now().isoformat(),
            "checks_total": 0,
            "checks_today": 0,
            "checks_date": str(date.today()),
            "wa_connected": False,
            "banned": False,
        }
        save_data(db)
    else:
        db["users"][uid]["name"]  = name
        db["users"][uid]["username"] = uname

def is_banned(uid) -> bool:
    return str(uid) in db.get("banned", [])

def is_open_mode() -> bool:
    return db.get("settings", {}).get("open_mode", True)

def is_approved(uid) -> bool:
    if is_open_mode(): return True
    return str(uid) in db.get("approved", [])

def global_wa_on() -> bool:
    return db.get("global_wa", {}).get("enabled", True)

def global_wa_connected() -> bool:
    return db.get("global_wa", {}).get("connected", False)

def get_wa_uid(uid: str) -> str:
    if global_wa_on():
        return ADMIN_WA_UID
    return f"user_{uid}"

def check_daily_limit(uid: str) -> tuple:
    """Returns (can_check, remaining)"""
    users = db.get("users", {})
    u     = users.get(uid, {})
    today = str(date.today())

    # Reset today counter if new day
    if u.get("checks_date") != today:
        u["checks_today"] = 0
        u["checks_date"]  = today
        save_data(db)

    # Per-user limit (overrides global if set)
    user_limit = u.get("custom_limit", -1)
    if user_limit == -1:
        # Use global limit
        global_limit = db.get("settings", {}).get("daily_limit", 0)
        if global_limit == 0:
            return True, -1  # unlimited
        limit = global_limit
    elif user_limit == 0:
        return True, -1  # this user unlimited
    else:
        limit = user_limit

    done = u.get("checks_today", 0)
    return done < limit, max(0, limit - done)

def add_checks(uid: str, count: int):
    u     = db["users"].get(uid, {})
    today = str(date.today())
    if u.get("checks_date") != today:
        u["checks_today"] = 0
        u["checks_date"]  = today
    u["checks_today"]  = u.get("checks_today", 0) + count
    u["checks_total"]  = u.get("checks_total", 0) + count
    save_data(db)


# ═══════════════════════════════════════
# ─── Baileys ───
# ═══════════════════════════════════════

def baileys_req(method: str, path: str, body=None) -> dict:
    url  = f"{BAILEYS_URL}{path}"
    data = json.dumps(body).encode() if body else None
    try:
        req = urllib.request.Request(
            url, data=data,
            headers={"Content-Type": "application/json"},
            method=method
        )
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read().decode())
    except Exception as e:
        logger.error(f"Baileys [{path}]: {e}")
        return {}

async def wa_state(wa_uid: str) -> bool:
    loop   = asyncio.get_event_loop()
    result = await loop.run_in_executor(
        None, lambda: baileys_req("GET", f"/status?userId={wa_uid}")
    )
    return result.get("connected", False)

async def wa_pair(phone: str, wa_uid: str) -> str:
    loop   = asyncio.get_event_loop()
    digits = re.sub(r"\D", "", phone)
    await loop.run_in_executor(None, lambda: baileys_req("POST", "/start", {"userId": wa_uid}))
    await asyncio.sleep(3)
    result = await loop.run_in_executor(
        None, lambda: baileys_req("POST", "/pair", {"phone": digits, "userId": wa_uid})
    )
    return result.get("code", "")

async def wa_check(numbers: list, wa_uid: str) -> dict:
    loop    = asyncio.get_event_loop()
    results = {}
    for i in range(0, len(numbers), 50):
        batch = numbers[i:i+50]
        try:
            res = await loop.run_in_executor(
                None,
                lambda b=batch: baileys_req("POST", "/check", {"numbers": b, "userId": wa_uid})
            )
            results.update(res.get("results", {}))
        except Exception as e:
            logger.error(f"Check error: {e}")
            for n in batch: results[n] = None
        await asyncio.sleep(0.5)
    return results

async def wa_disconnect(wa_uid: str):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, lambda: baileys_req("POST", "/disconnect", {"userId": wa_uid}))


# ═══════════════════════════════════════
# ─── WA Monitor ───
# ═══════════════════════════════════════

async def wa_monitor(app):
    """WA connection monitor — disconnect হলে notify করো"""
    logger.info("📡 WA Monitor started")
    fail_counts = {}

    while True:
        await asyncio.sleep(30)
        try:
            # Admin global WA monitor
            if global_wa_connected():
                connected = await wa_state(ADMIN_WA_UID)
                if not connected:
                    fail_counts["admin"] = fail_counts.get("admin", 0) + 1
                    if fail_counts["admin"] >= 3:
                        fail_counts.pop("admin", None)
                        db["global_wa"]["connected"] = False
                        save_data(db)
                        try:
                            await app.bot.send_message(
                                OWNER_ID,
                                "⚠️ *Global WhatsApp Disconnected!*\n\n"
                                "Admin WA disconnect হয়েছে।\n"
                                "⚙️ Admin Panel → Global WA → Connect করুন।",
                                parse_mode="Markdown"
                            )
                        except: pass
                else:
                    fail_counts.pop("admin", None)

            # User WA monitor
            if not global_wa_on():
                for uid, u in list(db["users"].items()):
                    if not u.get("wa_connected"): continue
                    wa_uid    = f"user_{uid}"
                    connected = await wa_state(wa_uid)
                    if not connected:
                        fail_counts[uid] = fail_counts.get(uid, 0) + 1
                        if fail_counts[uid] >= 3:
                            fail_counts.pop(uid, None)
                            db["users"][uid]["wa_connected"] = False
                            save_data(db)
                            try:
                                await app.bot.send_message(
                                    int(uid),
                                    "⚠️ *WhatsApp Disconnected!*\n\n"
                                    "তোমার WhatsApp disconnect হয়েছে।\n"
                                    "📱 WA Connect button চাপো।",
                                    parse_mode="Markdown",
                                    reply_markup=main_kb(int(uid))
                                )
                            except: pass
                    else:
                        fail_counts.pop(uid, None)
        except Exception as e:
            logger.error(f"WA Monitor error: {e}")


# ═══════════════════════════════════════
# ─── Keyboards ───
# ═══════════════════════════════════════

def main_kb(uid: int):
    kb = [
        [KeyboardButton("🔍 Check Numbers"), KeyboardButton("📁 Upload File")],
        [KeyboardButton("📱 WA Connect"),    KeyboardButton("📊 My Status")],
        [KeyboardButton("💬 Support")],
    ]
    if uid == OWNER_ID:
        kb.append([KeyboardButton("⚙️ Admin Panel")])
    return ReplyKeyboardMarkup(kb, resize_keyboard=True)

def admin_menu():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👥 Users",        callback_data="adm_users:0"),
         InlineKeyboardButton("🚫 Banned",       callback_data="adm_banned:0")],
        [InlineKeyboardButton("📢 Broadcast",    callback_data="adm_broadcast"),
         InlineKeyboardButton("✉️ Msg User",     callback_data="adm_msg_user")],
        [InlineKeyboardButton("📱 Global WA",    callback_data="adm_global_wa"),
         InlineKeyboardButton("📊 Stats",        callback_data="adm_stats")],
        [InlineKeyboardButton("⚙️ Global Limit", callback_data="adm_limit"),
         InlineKeyboardButton("👤 User Limit",    callback_data="adm_user_limit")],
        [InlineKeyboardButton("📥 Export Users", callback_data="adm_export")],
        [InlineKeyboardButton("🔙 Close",          callback_data="adm_close")],
    ])


# ═══════════════════════════════════════
# ─── Core Check Function ───
# ═══════════════════════════════════════

async def do_check(update: Update, context: ContextTypes.DEFAULT_TYPE, numbers: list):
    uid    = str(update.effective_user.id)
    wa_uid = get_wa_uid(uid)

    # Check WA connected
    connected = await wa_state(wa_uid)
    if not connected:
        if global_wa_on():
            return await update.message.reply_text(
                "❌ Global WhatsApp এখনো connect হয়নি। Admin এর সাথে যোগাযোগ করুন।"
            )
        return await update.message.reply_text(
            "❌ WhatsApp connected নেই!\n📱 WA Connect button চাপুন।",
            reply_markup=main_kb(int(uid))
        )

    # Daily limit check
    can_check, remaining = check_daily_limit(uid)
    if not can_check:
        limit = db.get("settings", {}).get("daily_limit", 0)
        return await update.message.reply_text(
            f"⛔ দৈনিক limit শেষ!\n\n"
            f"আজকের limit: *{limit}* numbers\n"
            f"কাল আবার try করুন।",
            parse_mode="Markdown"
        )

    # Trim to limit if needed
    limit = db.get("settings", {}).get("daily_limit", 0)
    if limit > 0 and len(numbers) > remaining:
        numbers = numbers[:remaining]
        await update.message.reply_text(f"⚠️ Limit অনুযায়ী শুধু *{len(numbers)}* নম্বর check হবে।", parse_mode="Markdown")

    loading = await update.message.reply_text(
        f"⏳ *{len(numbers)}* নম্বর check করছি...", parse_mode="Markdown"
    )

    results    = await wa_check(numbers, wa_uid)
    registered = [n for n in numbers if results.get(n) is True]
    fresh      = [n for n in numbers if results.get(n) is False]
    failed     = [n for n in numbers if results.get(n) is None]
    ts         = datetime.now().strftime("%Y%m%d_%H%M%S")

    add_checks(uid, len(numbers))

    await loading.edit_text(
        f"✅ *Check Complete!*\n\n"
        f"📊 Total: *{len(numbers)}*\n"
        f"✅ WA Registered: *{len(registered)}*\n"
        f"✨ Fresh (no WA): *{len(fresh)}*\n"
        f"⚠️ Failed: *{len(failed)}*",
        parse_mode="Markdown"
    )

    if registered:
        await update.message.reply_document(
            document="\n".join(registered).encode(),
            filename=f"registered_{ts}.txt",
            caption=f"✅ *WA Registered* — {len(registered)} numbers",
            parse_mode="Markdown"
        )
    if fresh:
        await update.message.reply_document(
            document="\n".join(fresh).encode(),
            filename=f"fresh_{ts}.txt",
            caption=f"✨ *Fresh Numbers* — {len(fresh)} numbers",
            parse_mode="Markdown"
        )
    if failed:
        await update.message.reply_document(
            document="\n".join(failed).encode(),
            filename=f"failed_{ts}.txt",
            caption=f"⚠️ *Failed* — {len(failed)} numbers",
            parse_mode="Markdown"
        )


# ═══════════════════════════════════════
# ─── User Handlers ───
# ═══════════════════════════════════════

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if is_banned(uid):
        return await update.message.reply_text("🚫 You are banned from this bot.")
    register_user(update)
    if not is_approved(uid) and uid != OWNER_ID:
        return await update.message.reply_text(
            "🔒 *This bot is in closed mode.*\n\n"
            "Access এর জন্য admin এর সাথে যোগাযোগ করুন:\n"
            f"{SUPPORT_USER}",
            parse_mode="Markdown"
        )
    wa_uid    = get_wa_uid(str(uid))
    connected = await wa_state(wa_uid)
    status    = "🟢 Connected" if connected else "🔴 Not Connected"
    await update.message.reply_text(
        f"👋 *WhatsApp Number Checker Bot*\n\n"
        f"📱 WA Status: {status}\n\n"
        f"*কীভাবে use করবেন:*\n"
        f"• Text এ নম্বর লিখুন (যেকোনো format)\n"
        f"• 📁 File upload করুন (.txt/.xlsx/.xls)\n"
        f"• Result দুইটা file এ পাবেন",
        parse_mode="Markdown",
        reply_markup=main_kb(uid)
    )

async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    text = update.message.text.strip()
    if is_banned(uid): return
    register_user(update)
    sess = context.user_data

    # ── Admin states ──
    if uid == OWNER_ID:
        state = sess.get("state")

        if state == "broadcast":
            sess.clear()
            users  = db.get("users", {})
            sent   = 0
            failed = 0
            loading = await update.message.reply_text(f"📢 Broadcasting to {len(users)} users...")
            for u_id in users:
                try:
                    await context.bot.send_message(int(u_id), text, parse_mode="Markdown")
                    sent += 1
                except: failed += 1
                await asyncio.sleep(0.05)
            await loading.edit_text(f"✅ Broadcast done!\n✅ Sent: {sent}\n❌ Failed: {failed}")
            return

        if state == "msg_user":
            sess.clear()
            parts = text.split("\n", 1)
            if len(parts) < 2:
                return await update.message.reply_text("❌ Format:\n`USER_ID\nMessage`", parse_mode="Markdown")
            target_id, msg = parts[0].strip(), parts[1].strip()
            try:
                await context.bot.send_message(int(target_id), f"📩 *Admin Message:*\n\n{msg}", parse_mode="Markdown")
                await update.message.reply_text("✅ Message sent!")
            except Exception as e:
                await update.message.reply_text(f"❌ Failed: {e}")
            return

        if state == "set_limit":
            sess.clear()
            try:
                limit = int(text.strip())
                db.setdefault("settings", {})["daily_limit"] = limit
                save_data(db)
                await update.message.reply_text(
                    f"✅ Daily limit set to *{limit}* numbers\n{'(0 = unlimited)' if limit == 0 else ''}",
                    parse_mode="Markdown"
                )
            except:
                await update.message.reply_text("❌ সংখ্যা দিন।")
            return

        if state == "ban_user":
            sess.clear()
            target = text.strip()
            if target not in db.get("banned", []):
                db.setdefault("banned", []).append(target)
                save_data(db)
            await update.message.reply_text(f"🚫 User `{target}` banned.", parse_mode="Markdown")
            return

        if state == "unban_user":
            sess.clear()
            target = text.strip()
            if target in db.get("banned", []):
                db["banned"].remove(target)
                save_data(db)
            await update.message.reply_text(f"✅ User `{target}` unbanned.", parse_mode="Markdown")
            return

    # ── WA Connect flow ──
    if sess.get("state") == "wa_connect_phone":
        sess.clear()
        phone   = re.sub(r"\D", "", text)
        if len(phone) < 8:
            return await update.message.reply_text("❌ নম্বর ভুল!")
        wa_uid  = get_wa_uid(str(uid))
        loading = await update.message.reply_text("⏳ Connecting WhatsApp...")
        try:
            code = await wa_pair(phone, wa_uid)
            await loading.delete()
            if not code:
                return await update.message.reply_text("❌ Pairing code পাওয়া যায়নি। আবার try করুন।")
            await update.message.reply_text(
                f"📱 *Pairing Code:*\n\n`{code}`\n\n"
                f"WhatsApp → Linked Devices → Link a Device → Enter code\n\n"
                f"✅ Connected হলে `/start` দিন।",
                parse_mode="Markdown"
            )
            # Verify connection
            async def verify():
                for _ in range(20):
                    await asyncio.sleep(15)
                    if await wa_state(wa_uid):
                        if uid == OWNER_ID:
                            db["global_wa"]["connected"] = True
                            save_data(db)
                        else:
                            db["users"][str(uid)]["wa_connected"] = True
                            save_data(db)
                        try:
                            await context.bot.send_message(
                                uid,
                                "✅ *WhatsApp Connected!*\n\nএখন নম্বর check করতে পারবেন।",
                                parse_mode="Markdown",
                                reply_markup=main_kb(uid)
                            )
                        except: pass
                        return
            asyncio.create_task(verify())
        except Exception as e:
            await loading.delete()
            await update.message.reply_text(f"❌ Error: {e}")
        return

    # ── Button handlers ──
    if text == "⚙️ Admin Panel" and uid == OWNER_ID:
        await update.message.reply_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())
        return

    if text == "📱 WA Connect":
        context.user_data["state"] = "wa_connect_phone"
        await update.message.reply_text(
            "📱 *WhatsApp Connect*\n\nআপনার WhatsApp নম্বর দিন (country code সহ):\nExample: `8801712345678`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )
        return

    if text == "🔍 Check Numbers":
        await update.message.reply_text(
            "📝 নম্বর লিখুন বা paste করুন:\n\n"
            "যেকোনো format এ দিতে পারেন:\n"
            "`8801712345678\n8801987654321`\n\n"
            "অথবা comma/space দিয়েও দিতে পারেন।",
            parse_mode="Markdown"
        )
        context.user_data["state"] = "waiting_numbers"
        return

    if text == "📁 Upload File":
        await update.message.reply_text(
            "📁 `.txt`, `.xlsx` বা `.xls` file পাঠান।\nFile এ নম্বর থাকলেই হবে।"
        )
        return

    if text == "📊 My Status":
        u     = get_user(str(uid))
        limit = db.get("settings", {}).get("daily_limit", 0)
        _, rem = check_daily_limit(str(uid))
        wa_uid    = get_wa_uid(str(uid))
        connected = await wa_state(wa_uid)
        await update.message.reply_text(
            f"📊 *Your Status*\n\n"
            f"👤 Name: {u.get('name', 'N/A')}\n"
            f"🆔 ID: `{uid}`\n"
            f"📱 WA: {'🟢 Connected' if connected else '🔴 Disconnected'}\n"
            f"📊 Total Checks: *{u.get('checks_total', 0)}*\n"
            f"📅 Today: *{u.get('checks_today', 0)}*\n"
            f"⚡ Daily Limit: *{'Unlimited' if limit == 0 else f'{limit} (remaining: {rem})'}*\n"
            f"📅 Joined: {u.get('joined', 'N/A')[:10]}",
            parse_mode="Markdown"
        )
        return

    if text == "💬 Support":
        await update.message.reply_text(
            f"💬 *Support*\n\nযোগাযোগ করুন: {SUPPORT_USER}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("💬 Contact Support", url=f"https://t.me/{SUPPORT_USER.lstrip('@')}")
            ]])
        )
        return

    # ── Admin state check ──
    if sess.get("state"):
        return

    # ── Number input (text) ──
    if sess.get("state") == "waiting_numbers" or re.search(r'\d{7,15}', text):
        sess.pop("state", None)
        numbers = re.findall(r'\d{7,15}', text)
        numbers = list(dict.fromkeys(numbers))
        if not numbers:
            return await update.message.reply_text("❌ কোনো valid নম্বর পাওয়া যায়নি।")
        await do_check(update, context, numbers)
        return

async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if is_banned(uid): return
    register_user(update)
    if not is_approved(uid) and uid != OWNER_ID:
        return await update.message.reply_text("🔒 Access denied. Contact admin.")

    doc = update.message.document
    if not doc: return
    fname = doc.file_name.lower()
    if not (fname.endswith(".txt") or fname.endswith(".xlsx") or fname.endswith(".xls")):
        return await update.message.reply_text("❌ শুধু `.txt`, `.xlsx` বা `.xls` file পাঠান।")

    loading  = await update.message.reply_text("⏳ File পড়ছি...")
    tmp_path = f"/tmp/{doc.file_id}_{doc.file_name}"
    numbers  = []
    try:
        file = await context.bot.get_file(doc.file_id)
        await file.download_to_drive(tmp_path)

        if fname.endswith(".txt"):
            with open(tmp_path, "r", encoding="utf-8", errors="ignore") as f:
                numbers = re.findall(r'\d{7,15}', f.read())
        elif fname.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path, read_only=True, data_only=True)
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            d = re.sub(r"\D", "", str(cell))
                            if 7 <= len(d) <= 15: numbers.append(d)
        elif fname.endswith(".xls"):
            import pandas as pd
            df = pd.read_excel(tmp_path, engine="xlrd", dtype=str, header=None)
            for col in df.columns:
                for val in df[col].dropna():
                    d = re.sub(r"\D", "", str(val))
                    if 7 <= len(d) <= 15: numbers.append(d)
        os.remove(tmp_path)
    except Exception as e:
        await loading.delete()
        return await update.message.reply_text(f"❌ File read error: {e}")

    numbers = list(dict.fromkeys(numbers))
    await loading.delete()
    if not numbers:
        return await update.message.reply_text("❌ File এ কোনো নম্বর পাওয়া যায়নি।")

    await do_check(update, context, numbers)


# ═══════════════════════════════════════
# ─── Admin Callbacks ───
# ═══════════════════════════════════════

async def cb_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    uid   = update.effective_user.id
    data  = query.data
    try: await query.answer()
    except: pass

    # Cancel
    if data == "cancel":
        context.user_data.clear()
        try:
            await query.edit_message_text("❌ Cancelled")
        except: pass
        return

    if data == "adm_close":
        try: await query.delete_message()
        except: pass
        return

    if uid != OWNER_ID: return

    # ── Stats ──
    if data == "adm_stats":
        users   = db.get("users", {})
        banned  = db.get("banned", [])
        limit   = db.get("settings", {}).get("daily_limit", 0)
        gwa     = db.get("global_wa", {})
        total_checks = sum(u.get("checks_total", 0) for u in users.values())
        await query.edit_message_text(
            f"📊 *Bot Statistics*\n\n"
            f"👥 Total Users: *{len(users)}*\n"
            f"🔑 Mode: *{'🔓 Open' if is_open_mode() else '🔒 Closed'}*\n"
            f"✅ Approved: *{len(db.get('approved', []))}*\n"
            f"🚫 Banned: *{len(banned)}*\n"
            f"🔢 Total Checks: *{total_checks}*\n"
            f"⚡ Daily Limit: *{'Unlimited' if limit == 0 else limit}*\n"
            f"📱 Global WA: *{'✅ ON' if gwa.get('enabled') else '❌ OFF'}* | "
            f"{'🟢 Connected' if gwa.get('connected') else '🔴 Disconnected'}",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 Back", callback_data="adm_back")]])
        )

    # ── Users list ──
    elif data.startswith("adm_users:"):
        page  = int(data.split(":")[1])
        users = list(db.get("users", {}).items())
        total = len(users)
        start = page * USERS_PER_PAGE
        end   = start + USERS_PER_PAGE
        chunk = users[start:end]

        text = f"👥 *Users* (Page {page+1}/{max(1,(total+USERS_PER_PAGE-1)//USERS_PER_PAGE)})\n\n"
        for u_id, u in chunk:
            banned_mark = "🚫" if u_id in db.get("banned", []) else ""
            text += f"{banned_mark}👤 {u.get('name','N/A')} | `{u_id}`\n"
            text += f"   📊 {u.get('checks_total',0)} checks\n\n"

        buttons = []
        nav = []
        if page > 0:
            nav.append(InlineKeyboardButton("◀️", callback_data=f"adm_users:{page-1}"))
        if end < total:
            nav.append(InlineKeyboardButton("▶️", callback_data=f"adm_users:{page+1}"))
        if nav: buttons.append(nav)
        buttons.append([
            InlineKeyboardButton("🚫 Ban User",   callback_data="adm_do_ban"),
            InlineKeyboardButton("✅ Unban User", callback_data="adm_do_unban"),
        ])
        buttons.append([InlineKeyboardButton("🔙 Back", callback_data="adm_back")])
        await query.edit_message_text(text, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(buttons))

    # ── Banned list ──
    elif data.startswith("adm_banned"):
        banned = db.get("banned", [])
        text   = f"🚫 *Banned Users* ({len(banned)})\n\n"
        for b in banned[:20]:
            u = db["users"].get(b, {})
            text += f"• `{b}` — {u.get('name','Unknown')}\n"
        await query.edit_message_text(text or "No banned users.", parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ Unban", callback_data="adm_do_unban")],
                [InlineKeyboardButton("🔙 Back", callback_data="adm_back")]
            ]))

    # ── Ban/Unban ──
    elif data == "adm_do_ban":
        context.user_data["state"] = "ban_user"
        await query.edit_message_text(
            "🚫 Ban করতে User ID দিন:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )
    elif data == "adm_do_unban":
        context.user_data["state"] = "unban_user"
        await query.edit_message_text(
            "✅ Unban করতে User ID দিন:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    # ── Broadcast ──
    elif data == "adm_broadcast":
        context.user_data["state"] = "broadcast"
        await query.edit_message_text(
            "📢 *Broadcast Message*\n\nসকল user কে পাঠানোর message লিখুন:\n(Markdown supported)",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    # ── Message User ──
    elif data == "adm_msg_user":
        context.user_data["state"] = "msg_user"
        await query.edit_message_text(
            "✉️ *Message User*\n\nFormat:\n`USER_ID\nMessage text`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    # ── Mode Toggle ──
    elif data == "adm_mode_toggle":
        current = db.get("settings", {}).get("open_mode", True)
        db.setdefault("settings", {})["open_mode"] = not current
        save_data(db)
        mode = "🔓 Open (সবাই use করতে পারবে)" if not current else "🔒 Closed (শুধু approved)"
        await query.answer(f"Mode: {mode}", show_alert=True)
        await query.edit_message_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())

    # ── Approve ──
    elif data == "adm_approve":
        context.user_data["state"] = "approve_user"
        mode = "🔓 Open" if is_open_mode() else "🔒 Closed"
        approved_count = len(db.get("approved", []))
        await query.edit_message_text(
            f"✅ *Approve User*\n\n"
            f"Current Mode: *{mode}*\n"
            f"Approved Users: *{approved_count}*\n\n"
            f"Approve করতে User ID দিন:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    elif data == "adm_unapprove":
        context.user_data["state"] = "unapprove_user"
        approved = db.get("approved", [])
        text_list = "\n".join([f"• `{a}`" for a in approved[:20]]) or "কেউ নেই"
        await query.edit_message_text(
            f"❌ *Unapprove User*\n\nApproved list:\n{text_list}\n\nUnapprove করতে User ID দিন:",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    # ── User Limit ──
    elif data == "adm_user_limit":
        context.user_data["state"] = "set_user_limit"
        await query.edit_message_text(
            "👤 *User Limit Set*\n\n"
            "Format: `USER_ID LIMIT`\n\n"
            "Examples:\n"
            "`123456789 500` → 500/day\n"
            "`123456789 0` → Unlimited\n"
            "`123456789 -1` → Use Global limit",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    # ── Daily Limit ──
    elif data == "adm_limit":
        current = db.get("settings", {}).get("daily_limit", 0)
        context.user_data["state"] = "set_limit"
        await query.edit_message_text(
            f"⚡ *Daily Limit*\n\nCurrent: *{'Unlimited' if current == 0 else current}*\n\n"
            f"নতুন limit দিন (0 = unlimited):",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    # ── Global WA ──
    elif data == "adm_global_wa":
        gwa       = db.get("global_wa", {})
        enabled   = gwa.get("enabled", True)
        connected = gwa.get("connected", False)
        await query.edit_message_text(
            f"📱 *Global WhatsApp*\n\n"
            f"Mode: *{'🌍 Global (Admin WA)' if enabled else '👤 Per-User WA'}*\n"
            f"Status: *{'🟢 Connected' if connected else '🔴 Disconnected'}*\n\n"
            f"Global ON → সব user admin এর WA দিয়ে check করবে\n"
            f"Global OFF → প্রতিটা user নিজে WA connect করবে",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton(
                    f"{'🔴 Disable Global' if enabled else '🟢 Enable Global'}",
                    callback_data="adm_gwa_toggle"
                )],
                [InlineKeyboardButton("📱 Connect Admin WA", callback_data="adm_gwa_connect"),
                 InlineKeyboardButton("🔌 Disconnect",       callback_data="adm_gwa_disconnect")],
                [InlineKeyboardButton("🔙 Back", callback_data="adm_back")],
            ])
        )

    elif data == "adm_gwa_toggle":
        db["global_wa"]["enabled"] = not db["global_wa"].get("enabled", True)
        save_data(db)
        await query.answer(f"Global WA {'ON' if db['global_wa']['enabled'] else 'OFF'}", show_alert=True)
        await cb_handler(update, context)  # refresh

    elif data == "adm_gwa_connect":
        context.user_data["state"] = "wa_connect_phone"
        await query.edit_message_text(
            "📱 Admin WhatsApp নম্বর দিন:\nExample: `8801712345678`",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )

    elif data == "adm_gwa_disconnect":
        await wa_disconnect(ADMIN_WA_UID)
        db["global_wa"]["connected"] = False
        save_data(db)
        await query.answer("🔌 Disconnected", show_alert=True)
        # refresh
        query.data = "adm_global_wa"
        await cb_handler(update, context)

    elif data == "adm_export":
        users  = db.get("users", {})
        banned = db.get("banned", [])
        approved = db.get("approved", [])
        lines  = []
        lines.append("=" * 60)
        lines.append(f"WA CHECKER BOT - USER EXPORT")
        lines.append(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"Total Users: {len(users)}")
        lines.append("=" * 60)
        lines.append("")

        for uid, u in sorted(users.items(), key=lambda x: x[1].get("checks_total", 0), reverse=True):
            uname    = u.get("username", "")
            name     = u.get("name", "N/A")
            joined   = u.get("joined", "")[:10]
            total    = u.get("checks_total", 0)
            today    = u.get("checks_today", 0)
            climit   = u.get("custom_limit", -1)
            is_ban   = "🚫 BANNED" if uid in banned else ""
            is_app   = "✅ APPROVED" if uid in approved else ""
            limit_lbl = "Global" if climit == -1 else ("Unlimited" if climit == 0 else f"{climit}/day")

            lines.append(f"👤 Name     : {name}")
            lines.append(f"🆔 User ID  : {uid}")
            lines.append(f"📛 Username : @{uname}" if uname else f"📛 Username : (not set)")
            lines.append(f"📅 Joined   : {joined}")
            lines.append(f"📊 Total    : {total} checks")
            lines.append(f"📅 Today    : {today} checks")
            lines.append(f"⚡ Limit    : {limit_lbl}")
            if is_ban:  lines.append(f"Status     : {is_ban}")
            if is_app:  lines.append(f"Status     : {is_app}")
            lines.append("-" * 40)

        file_content = "\n".join(lines)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        await context.bot.send_document(
            chat_id=query.from_user.id,
            document=file_content.encode("utf-8"),
            filename=f"users_export_{ts}.txt",
            caption=f"📥 *User Export*\n\n👥 Total: *{len(users)}* users",
            parse_mode="Markdown"
        )
        await query.answer("✅ File sent!", show_alert=True)

    elif data == "adm_back":
        await query.edit_message_text("⚙️ *Admin Panel*", parse_mode="Markdown", reply_markup=admin_menu())


# ═══════════════════════════════════════
# ─── Main ───
# ═══════════════════════════════════════

def main():
    logger.info("🔍 WA Checker Bot Starting...")
    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CallbackQueryHandler(cb_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    async def post_init(application):
        asyncio.create_task(wa_monitor(application))

    app.post_init = post_init
    app.run_polling(allowed_updates=["message", "callback_query"])

if __name__ == "__main__":
    main()
